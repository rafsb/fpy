from openpyxl import load_workbook, styles, Workbook

from utils.basic_traits import class_t
from utils.palette import palette_t

fmt_date = styles.NamedStyle(name='fmt_date')
fmt_date.number_format = 'dd/mm/yyyy'
fmt_date.alignment = styles.Alignment(horizontal='center', vertical='center')
fmt_date.font = styles.Font(size=9)

fmt_number = styles.NamedStyle(name='fmt_number')
fmt_number.number_format = '#,##0.00;(#,##0.00)'
fmt_number.font = styles.Font(size=9)

fmt_text = styles.NamedStyle(name='fmt_text')
fmt_text.number_format = '@'
fmt_text.font = styles.Font(size=9)

fmt_percentage = styles.NamedStyle(name='fmt_percentage')
fmt_percentage.number_format = '0.00%'
fmt_percentage.font = styles.Font(size=9)

fmt_border = styles.Border(
    left=styles.Side(border_style='thin', color='D2D2D2')
    , right=styles.Side(border_style='thin', color='D2D2D2')
    , top=styles.Side(border_style='thin', color='D2D2D2')
    , bottom=styles.Side(border_style='thin', color='D2D2D2')
)

numericals = [
    ...
]

datetimes = [
    ...
]


class sheets(class_t):

    @staticmethod
    def color_gradient(color, grad=1, theme='light'):
        """
        Generates a gradient color based on an input color.

        This function takes a base color in hexadecimal format and creates a gradient
        version of it. The gradient can be either lighter or darker depending on the
        'theme' parameter.

        Args:
        color (str): The base color in hexadecimal format (e.g., '#RRGGBB')
        grad (float): A value between 0 and 1 that determines the strength of the gradient
        theme (str): Either 'light' or 'dark', determines if the gradient lightens or darkens the color

        Returns:
        str: A new color in hexadecimal format representing the gradient

        The function works by:
        1. Converting the hex color to RGB values
        2. Adjusting these RGB values based on the gradient and theme
        3. Converting the adjusted RGB values back to a hex color string
        4. Ensuring the resulting color values are within the valid range (0-255)
        """

        # Convert hex to RGB and apply gradient
        if len(color.lstrip('#')) == 6:
            if theme == 'light':
                rgb = [int(color.lstrip('#')[i:i + 2], 16) + ((255 - int(color.lstrip('#')[i:i + 2], 16)) * (1 - grad)) for i in (0, 2, 4)]
            else:
                rgb = [int(color.lstrip('#')[i:i + 2], 16) * grad for i in (0, 2, 4)]
        else:
            rgb = [0, 0, 0]

        # Convert back to hex, ensuring values are within 0-255 range
        return '#{:02x}{:02x}{:02x}'.format(
            min(255, max(0, int(rgb[0]))),
            min(255, max(0, int(rgb[1]))),
            min(255, max(0, int(rgb[2])))
        )

    def assign(self, sheet, row, column, value, style=None, fill=None, font=None, border=True, align=None):
        """
        Assigns properties to a cell in an Excel worksheet.

        This function sets various attributes of a cell including its value, style,
        border, fill color, alignment, and font properties.

        Args:
        cell (openpyxl.cell.Cell): The cell to modify
        value: The value to be assigned to the cell
        style: The style to be applied to the cell (default is fmt_text if not provided)
        fill (str): The fill color for the cell (in hex format)
        font (str): The font color for the cell (in hex format)
        border (bool): Whether to apply a border to the cell (default True)
        align (str): The horizontal alignment for the cell content

        Returns:
        openpyxl.cell.Cell: The modified cell object
        """
        if sheet not in self.wb.sheetnames:
            self.wb.create_sheet(sheet)
        sheet = self.wb[sheet]
        cell = sheet.cell(row=row, column=column)
        cell.value = value
        cell.style = style if style else fmt_text
        if border:
            cell.border = fmt_border
        if fill:
            cell.fill = styles.PatternFill(start_color=fill.lstrip('#').upper(), fill_type='solid')
        if align:
            cell.alignment = styles.Alignment(horizontal=align, vertical='center')

        if not value:
            cell.font = styles.Font(color=palette_t.get('DISABLED').lstrip('#').upper(), size=9)
        elif font:
            cell.font = styles.Font(color=font.lstrip('#').upper(), size=9)

        else:
            cell.font = styles.Font(size=9)

        return cell

    def set_global_font_size(self, size=10):
        """
        Sets the font size globally for all cells in all sheets of a workbook.

        This function iterates through all sheets in the given workbook,
        and for each cell, it sets the font size to the specified value.
        If a cell already has a font, it preserves other font properties
        like name, bold, italic, and color while updating the size.
        If a cell doesn't have a font, it creates a new font with the specified size.

        Args:
        wb (openpyxl.Workbook): The workbook to modify
        size (int): The font size to set (default is 9)

        Returns:
        None
        """
        for sheet in self.wb:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.font:
                        cell.font = styles.Font(size=size, name=cell.font.name, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color)
                    else:
                        cell.font = styles.Font(size=size)

    def fill_sheet(self, sheet, data):
        """
        Fills an Excel sheet with data from a data object.

        This function does the following:
        1. Loads translations for column headers
        2. Fills the first row with translated column headers
        3. Fills subsequent rows with data from the data object
        4. Applies appropriate formatting to cells based on data type
        5. Adds an auto filter to the sheet

        Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The sheet to fill
        data (object): An object containing cols and items attributes

        Returns:
        None
        """

        cols = data[0].keys()

        # Fill header row
        for i, h in enumerate(cols, start=1):
            self.assign(sheet, 1, i, h, fill=palette_t.get('candy_gray'), align='center')

        # Fill data rows
        for i, item in enumerate(data, start=2):
            for j, h in enumerate(cols, start=1):
                try:
                    if h in (self.numericals or []): self.assign(sheet, i, j, item.get(h, '') if type(item) is dict else getattr(item, h, ''), style=fmt_number)
                    elif h in (self.datetimes or []): self.assign(sheet, i, j, item.get(h, '') if type(item) is dict else getattr(item, h, ''), style=fmt_date)
                    else: self.assign(sheet, i, j, item.get(h, '') if type(item) is dict else getattr(item, h, ''))
                except:
                    pass

        # Add auto filter
        self.wb[sheet].auto_filter.ref = self.wb[sheet].dimensions

    def load(self, path):
        self.wb = load_workbook(path)
        self.set_global_font_size(10)
        return self.wb

    def cast(self, tabs):
        self.wb = Workbook()
        if tabs and len(tabs):
            for tab in tabs:
                self.wb.create_sheet(tab)
        return self.wb

    def __init__(self, **args):
        super().__init__(**args)
        self.wb = Workbook()


def register(app, *args): pass
